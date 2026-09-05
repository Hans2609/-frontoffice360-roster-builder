"""
template_filler.py
Takes scraped player dicts + a school name, and writes them into a copy
of the FrontOffice360 blank roster template, matching the exact column
layout (First Name, Last Name, Pos, Jersey Number, ..., Fun Facts, ...).

Only the deterministic fields (available directly from the official
roster page) are filled here. Judgment fields -- Fun Facts, Academic
Interests, and any position label the scraper couldn't confidently map --
are intentionally left blank, with a companion prompt (see prompt_builder.py)
generated for finishing those in Claude.
"""

import openpyxl
from io import BytesIO

# Column positions match the FrontOffice360_Athlete_Roster_Template.xlsx
# "Roster Template" sheet header row exactly.
COL = {
    "first_name": 1,
    "last_name": 2,
    "pos": 3,
    "jersey_number": 4,
    "height": 7,
    "weight": 8,
    "status": 10,
    "college": 11,
    "hometown": 12,
    "high_school": 13,
    "entry_year": 14,
    "red_shirt": 15,
}


def fill_template(template_path: str, players: list, college_name: str) -> BytesIO:
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Roster Template"]

    # clear any example row(s) below the header before writing
    max_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    row_idx = 2
    for p in players:
        is_redshirt = bool(p.get("class_year") and p["class_year"].strip().lower().startswith("r-"))

        ws.cell(row=row_idx, column=COL["first_name"]).value = p.get("first_name")
        ws.cell(row=row_idx, column=COL["last_name"]).value = p.get("last_name")
        ws.cell(row=row_idx, column=COL["pos"]).value = p.get("position")
        ws.cell(row=row_idx, column=COL["jersey_number"]).value = p.get("jersey_number")
        ws.cell(row=row_idx, column=COL["height"]).value = p.get("height")
        ws.cell(row=row_idx, column=COL["weight"]).value = p.get("weight")
        ws.cell(row=row_idx, column=COL["status"]).value = "Active"
        ws.cell(row=row_idx, column=COL["college"]).value = college_name
        ws.cell(row=row_idx, column=COL["hometown"]).value = p.get("hometown")
        ws.cell(row=row_idx, column=COL["high_school"]).value = p.get("high_school")
        ws.cell(row=row_idx, column=COL["entry_year"]).value = p.get("entry_year_estimate")
        ws.cell(row=row_idx, column=COL["red_shirt"]).value = is_redshirt
        row_idx += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
